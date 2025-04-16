import datetime
import statistics
from turtle import pd
import cv2
import numpy as np
import dlib
from imutils import face_utils
import imutils
import os
import time
from datetime import datetime
# import pyhrv
from scipy import signal, sparse, stats
import scipy.signal
import matplotlib.pyplot as plt
import matplotlib.dates as dates
from scipy.sparse import diags
from scipy.sparse.linalg import spsolve
from BaselineRemoval import BaselineRemoval
from sklearn.preprocessing import MinMaxScaler
from pandas import DataFrame
from scipy.signal import find_peaks
# import biosppy
# import pyhrv.tools as tools
# from opensignalsreader import OpenSignalsReader

def read_numbers_exclude_first_last(file_path):
    
    with open(file_path, "r") as file:
        lines = file.readlines()
        
        # 排除第一行和最後一行
        if len(lines) > 2:
            lines = lines[1:-1]
        else:
            return ""
        
        # 將剩餘的行合併成一個字串，並去除行尾的換行符號
        result = "".join(line.strip() for line in lines)
        
    return result

def plot_data(r,標題):
    
    # 創建一個新的圖形
    # plt.figure()

    # 將列表轉換為 NumPy 陣列
    # r = np.array(r)

    # # 對時間欄位（第一列）進行 Min-Max 歸一化
    # time_column = r[:, 1]
    # min_time = np.min(time_column)
    # max_time = np.max(time_column)
    # normalized_time = (time_column - min_time) / (max_time - min_time)

    # # 用歸一化後的時間欄位替換原始時間欄位
    # r[:, 1] = normalized_time
    # 將列表轉換為 NumPy 陣列
    
    axes = plt.gca()
    axes.xaxis.set_major_locator(dates.DayLocator(interval=20))
    axes.yaxis.set_major_locator(dates.DayLocator(interval=10))
    #autoformat_xdate
    plt.gcf().autofmt_xdate()
    # 繪製 r 的圖形
    plt.plot(  r[:, 1],r[:, 0].astype(float),label='Values', color='blue')
    
    
    # plt.plot( r[:, 0])
    # 添加標題和標籤
    plt.title(標題)
    plt.xlabel('Index')
    plt.ylabel('Value')

    # 顯示圖例
    plt.legend()

    # 顯示圖形
    # plt.show()

def 兩圖合併(np1,np2,_title):
    #微軟正黑體
    plt.rc('font', family='Heiti TC')

    #圖像設定盡量大一點
    plt.figure(figsize=(20, 10))

    #plot 1:
    # xpoints = np1.array([0, 6])
    # ypoints = np1.array([0, 100])
    if np1 is not None :
        # 將當前的圖形窗口分成 1 行 2 列，並激活第 1 個子圖。
        plt.subplot(2, 1, 1)
        #plt.plot(xpoints,ypoints)
        plt.plot(  np1[:, 1],np1[:, 0].astype(float),label='Values', color='blue')
        plt.title(_title + "前")

    #plot 2:
    if np2 is not None :
        # 將當前的圖形窗口分成 1 行 2 列，並激活第 2 個子圖。
        plt.subplot(2, 1, 2)
        # plt.plot(x,y)
        plt.plot(  np2[:, 1],np2[:, 0].astype(float),label='Values', color='blue')
        plt.title(_title + "後")

    plt.suptitle(_title + "前後比較")
    plt.show()

#讀取檔案(有區間和第幾個timeWindows條件)
def read_from_file(file_path,timeWindow,第幾個timeWindow,區間_from,區間_to):
    #numbers是一個二維的陣列，第一個維度是數字，第二個維度是時間
    numbers = []

    _time = ""
    
    with open(file_path, 'r') as file:
        #將檔案中的所有數字讀取出來成為一個字串
        lines = file.readlines()
        numbers = []
        timeWindowCnt = 0 #time window
        #設定一個y軸的數字，按照底下的條件來加入    
        y軸 = 0
        跳過幾個點 = (第幾個timeWindow -1) * timeWindow
        _n = 0

        #將每個數字轉換為小數點後兩位的字串格式，並加入到 numbers(List) 中
        for line in lines:
            if _n < 跳過幾個點:
                _n += 1
                continue
            # 將每行的數字用逗號分隔
            data_tpm = line.split(",")
            # 並將字串轉換為浮點數，取到小數點後兩位，並加入到 numbers 中

            #只有數字在區間_from和區間_to之間才加入
            if float(data_tpm[1].strip()) > 區間_from and float(data_tpm[1].strip()) < 區間_to:
                # 將data_tpm「1]的時間加入到numbers的第二欄位中
                # numbers.append([float(data_tpm[1].strip()),data_tpm[0][3:5].strip()+data_tpm[0][6:14].strip()])
                #使用y軸的數字
                y軸 #不做任何處理.. 
                #使用資料的時間
                #numbers.append([float(data_tpm[1].strip()),datetime.strptime(data_tpm[0],"%H:%M:%S.%f")])
                # numbers.append([float(data_tpm[1].strip()),pd.to_datetime("20241203 " + data_tpm[0],format="%Y%m%d %H:%M:%S.%f")])
            else:
                continue
            
            y軸 += 1
            #將data_tpm「1]的時間加入到numbers的第二欄位中
            # numbers.append([float(data_tpm[1].strip()),data_tpm[0][3:5].strip()+data_tpm[0][6:14].strip()])
            #使用y軸的數字
            # y軸 #不做任何處理.. 
            # #使用資料的時間
            # #numbers.append([float(data_tpm[1].strip()),datetime.strptime(data_tpm[0],"%H:%M:%S.%f")])
            # numbers.append([float(data_tpm[1].strip()),pd.to_datetime("20241203 " + data_tpm[0],format="%Y%m%d %H:%M:%S.%f")]) 

            try:
                 numbers.append([float(data_tpm[1].strip()),y軸,datetime.strptime(data_tpm[0],"%H:%M:%S.%f")])
            except Exception as e:
                continue

            
            #計算timeWindow，若大於timeWindow則跳出迴圈
            timeWindowCnt += 1
            if timeWindowCnt > timeWindow-1:
                break
        return numbers
#讀取檔案(有區間和第幾個timeWindows條件)
def read_ALL_from_file(file_path):
    #numbers是一個二維的陣列，第一個維度是數字，第二個維度是時間
    numbers = []

    start_time = ""
    
    with open(file_path, 'r') as file:
        #將檔案中的所有數字讀取出來成為一個字串
        lines = file.readlines()
        numbers = []
        timeWindowCnt = 0 #time window
        #設定一個y軸的數字，按照底下的條件來加入    
        y軸 = 0
        time_diff = 0

        #將每個數字轉換為小數點後兩位的字串格式，並加入到 numbers(List) 中
        for line in lines:
            
            # 將每行的數字用逗號分隔
            data_tpm = line.split(",")
            # 並將字串轉換為浮點數，取到小數點後兩位，並加入到 numbers 中
            

            y軸 += 1
            #將data_tpm「1]的時間加入到numbers的第二欄位中
            # numbers.append([float(data_tpm[1].strip()),data_tpm[0][3:5].strip()+data_tpm[0][6:14].strip()])
            #使用y軸的數字
            # y軸 #不做任何處理.. 
            # #使用資料的時間
            # #numbers.append([float(data_tpm[1].strip()),datetime.strptime(data_tpm[0],"%H:%M:%S.%f")])
            # numbers.append([float(data_tpm[1].strip()),pd.to_datetime("20241203 " + data_tpm[0],format="%Y%m%d %H:%M:%S.%f")]) 

            #將data_tpm[0]轉換爲時間格式，並計算時間差
            thistime = datetime.strptime(data_tpm[0],"%H:%M:%S.%f")
            if start_time == "":
                start_time = thistime

            else:
                #計算時間差
                time_diff = thistime - start_time
                #轉換為秒
                time_diff = time_diff.total_seconds()
                # #將data_tpm[0]轉換爲時間格式，並計算時間差
                # numbers.append([float(data_tpm[1].strip()),y軸,thistime,time_diff])
                _time = thistime

            try:
                # numbers.append([float(data_tpm[1].strip()),y軸,datetime.strptime(data_tpm[0],"%H:%M:%S.%f")])
                numbers.append([float(data_tpm[1].strip()),y軸,time_diff])

            except Exception as e:
                continue

        return numbers   
    
def load_signals_NParray(self,frame,region):
        # 存放檔案的資料夾
        file_path_r = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_b.txt"
        file_path_g = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_b.txt"
        file_path_b = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_b.txt"

        #一次將檔案中的所有數字讀取出來成為一個字串，但第一行和最後一行的數字是不需要的
        numbers_list_r = read_from_file(file_path_r)
        numbers_list_g = read_from_file(file_path_g)
        numbers_list_b = read_from_file(file_path_b)
        
        tmpArray = np.zeros((10, 10, 3), np.uint8)

        #將numbers_list_r、numbers_list_g、numbers_list_b轉換為tmpArray
        for i in range(10):
            for j in range(10):
                tmpArray[i][j][0] = numbers_list_b[i][0]
                tmpArray[i][j][1] = numbers_list_g[i][0]
                tmpArray[i][j][2] = numbers_list_r[i][0]

        return tmpArray

def load_signals_LIST():
        # 存放檔案的資料夾
        file_path_r = "/Users/yangjames/Documents/python/rPPGData/signal/20241129/cheeck/20241129_b.txt"
        # file_path_r = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_r.txt"
        # file_path_r = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_g.txt"
        # file_path_g = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_b.txt"
        # file_path_b = "/Users/yangjames/Documents/python/rPPGData/signal/cheeck_20240924/20240924_b.txt"

        #一次將檔案中的所有數字讀取出來成為一個字串，但第一行和最後一行的數字是不需要的
        return read_from_file(file_path_r)

def butterworth_bandpass_filter(self, data, lowcut, highcut, fs, order=1):
        b, a = self.butterworth_bandpass(lowcut, highcut, fs, order=order)
        y = signal.lfilter(b, a, data)
        return y       

def smoothTriangle(data, degree=25):
        triangle=np.concatenate((np.arange(degree + 1), np.arange(degree)[::-1])) # up then down
        smoothed=[]
    
        for i in range(degree, len(data) - degree * 2):
            point=data[i:i + len(triangle)] * triangle
            smoothed.append(np.sum(point)/np.sum(triangle))
        # Handle boundaries
        smoothed=[smoothed[0]]*int(degree + degree/2) + smoothed
        while len(smoothed) < len(data):
            smoothed.append(smoothed[-1])
        return smoothed  

#基線校準(进行AsLS基线校正)
def baseline_als(y, lam=1e5, p=0.01, niter=10):
    L = len(y) 
    D = diags([1, -2, 1], [0, -1, -2], shape=(L, L-2)) 
    D = lam * D.dot(D.T) 
    w = np.ones(L) 
    for i in range(niter): 
        W = diags(w, 0, shape=(L, L)) 
        Z = W + D 
        z = spsolve(Z, w*y) 
        w = p * (y > z) + (1-p) * (y < z) 
    return z
#基線校準(主要呼叫)
def baseline_correction(input_data): 
    corrected_spectra = np.zeros_like(input_data) 
    for i in range(input_data.shape[0]): 
        baseline_values = baseline_als(input_data[i, :]) 
        corrected_spectra[i, :] = input_data[i, :] - baseline_values 
    return corrected_spectra

#基線校準(上一屆作法)
def detrend(data, axis=-1, type='linear', bp=0, overwrite_data=False):
    """
    Remove linear trend along axis from data.

    Parameters
    ----------
    data : array_like
        The input data.
    axis : int, optional
        The axis along which to detrend the data. By default this is the
        last axis (-1).
    type : {'linear', 'constant'}, optional
        The type of detrending. If ``type == 'linear'`` (default),
        the result of a linear least-squares fit to `data` is subtracted
        from `data`.
        If ``type == 'constant'``, only the mean of `data` is subtracted.
    bp : array_like of ints, optional
        A sequence of break points. If given, an individual linear fit is
        performed for each part of `data` between two break points.
        Break points are specified as indices into `data`. This parameter
        only has an effect when ``type == 'linear'``.
    overwrite_data : bool, optional
        If True, perform in place detrending and avoid a copy. Default is False

    Returns
    -------
    ret : ndarray
        The detrended input data.

    Examples
    --------
    >>> import numpy as np
    >>> from scipy import signal
    >>> rng = np.random.default_rng()
    >>> npoints = 1000
    >>> noise = rng.standard_normal(npoints)
    >>> x = 3 + 2*np.linspace(0, 1, npoints) + noise
    >>> (signal.detrend(x) - noise).max()
    0.06  # random

    """
    if type not in ['linear', 'l', 'constant', 'c']:
        raise ValueError("Trend type must be 'linear' or 'constant'.")
    data = np.asarray(data)
    dtype = data.dtype.char
    if dtype not in 'dfDF':
        dtype = 'd'
    if type in ['constant', 'c']:
        ret = data - np.mean(data, axis, keepdims=True)
        return ret
    else:
        dshape = data.shape
        N = dshape[axis]
        bp = np.sort(np.unique(np.r_[0, bp, N]))
        if np.any(bp > N):
            raise ValueError("Breakpoints must be less than length "
                             "of data along given axis.")
        Nreg = len(bp) - 1
        # Restructure data so that axis is along first dimension and
        #  all other dimensions are collapsed into second dimension
        rnk = len(dshape)
        if axis < 0:
            axis = axis + rnk
        newdims = np.r_[axis, 0:axis, axis + 1:rnk]
        newdata = np.reshape(np.transpose(data, tuple(newdims)),
                             (N, _prod(dshape) // N))
        if not overwrite_data:
            newdata = newdata.copy()  # make sure we have a copy
        if newdata.dtype.char not in 'dfDF':
            newdata = newdata.astype(dtype)
        # Find leastsq fit and remove it for each piece
        for m in range(Nreg):
            Npts = bp[m + 1] - bp[m]
            A = np.ones((Npts, 2), dtype)
            A[:, 0] = np.cast[dtype](np.arange(1, Npts + 1) * 1.0 / Npts)
            sl = slice(bp[m], bp[m + 1])
            coef, resids, rank, s = linalg.lstsq(A, newdata[sl])
            newdata[sl] = newdata[sl] - np.dot(A, coef)
        # Put data back in original shape.
        tdshape = np.take(dshape, newdims, 0)
        ret = np.reshape(newdata, tuple(tdshape))
        vals = list(range(1, rnk))
        olddims = vals[:axis] + [0] + vals[axis:]
        ret = np.transpose(ret, tuple(olddims))
        return ret

#歸一化用...
def Z_ScoreNormalization(x,mu,sigma):
        x = (x - mu) / sigma;
        return x;
##巴特濾波器
def butterworth_bandpass(lowcut, highcut, fs, order=1):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    b, a = signal.butter(order, [low, high], btype='band',analog='True')
    return b, a
    #order 是濾波器的階數，階數越大，濾波效果越好，但是計算量也會跟著變大。
    #所產生的濾波器參數 a 和 b 的長度，等於 order+1。
    #Wn 是正規化的截止頻率，介於 0 和 1 之間，當取樣頻率是 fs 時，所能處理的
    #最高頻率是 fs/2，所以如果實際的截止頻率是 f = 1000，那麼 Wn = f/(fs/2)。
    #function 是一個字串，function = 'low' 代表是低通濾波器，function = 'high' 代表是高通濾波。
    #fs=12,wn=f/(fs/2),如果截止頻率大於6,就高於正規化的截止頻率

def butterworth_bandpass_filter(data, lowcut, highcut, fs, order=1):
    b, a = butterworth_bandpass(lowcut, highcut, fs, order=order)
    y = signal.lfilter(b, a, data)
    return y

# 多项式拟合和基线校正
def baseline_correction(x, y_uniform, n):
    # 初始多项式拟合
    p0 = np.polyfit(x, y_uniform, n)  # 多项式拟合，返回多项式系数
    y_fit0 = np.polyval(p0, x)  # 计算拟合值
    r0 = y_uniform - y_fit0  # 残差
    dev0 = np.sqrt(np.sum((r0 - np.mean(r0)) ** 2) / len(r0))  # 计算残差

    # 峰值消除
    y_remove0 = y_uniform[y_uniform <= y_fit0]
    x_remove0 = x[np.where(y_uniform <= y_fit0)]

    # 初始化循环变量
    i = 0
    judge = 1
    dev = []

    while judge:
        # 多项式拟合
        p1 = np.polyfit(x_remove0, y_remove0, n)
        y_fit1 = np.polyval(p1, x_remove0)  # 计算拟合值
        r1 = y_remove0 - y_fit1  # 计算残差
        dev1 = np.sqrt(np.sum((r1 - np.mean(r1)) ** 2) / len(r1))  # 计算残差
        dev.append(dev1)

        # 判断残差变化
        if i == 0:
            judge = abs(dev[i] - dev0) / dev[i] > 0.05
        else:
            judge = abs((dev[i] - dev[i - 1]) / dev[i]) > 0.05

        # 光谱重建
        y_remove0[np.where(y_remove0 >= y_fit1)] = y_fit1[np.where(y_remove0 >= y_fit1)]

        i += 1

    # 最终基线和基线校正结果
    y_baseline = np.polyval(p1, x)  # 基线
    y_baseline_correction = y_uniform - y_baseline  # 基线校正后
    return y_baseline_correction
    # return y_baseline, y_baseline_correction

# 2024.12.01找到的使用bandpass Butterworth filter对信号数据进行滤波去噪
def bandPass_filter(signal):
    # # fs = 4000.0
    # fs = 15.0
    # # lowcut = 20.0 0.5以下
    # lowcut = 0.4
    # # highcut = 50.0 20-25
    # highcut = 25.0
    # nyqs = 0.5* fs
    # low = lowcut / nyqs
    # high = highcut/ nyqs
    low = 0.1
    high = 0.5
    
    # order=2 2-4
    order=2
    
    #一般數位訊號處理請使用 analog=False（預設），因為訊號通常是數位訊號
    b, a = scipy.signal.butter(order, [low, high], 'bandpass', analog = True) # b: 滤波器的分子系数向量，a: 滤波器的分母系数向量
    y = scipy.signal.filtfilt(b, a, signal, axis=0)  # 滤波函数
    return(y)
#計算心率
def BPM(bpm):
    bpm_plot = []
    bpm_plot.extend(bpm)        
    return bpm_plot
def calculate_bpm(ecg_signal, sampling_rate):
    """
    計算心率（BPM）。
    :param ecg_signal: np.array, ECG 信號數據
    :param sampling_rate: int, 取樣率（每秒樣本數）
    :return: float, 心率（BPM）
    """
    # 使用 find_peaks 函數找到 R 峰
    peaks, _ = find_peaks(ecg_signal, distance=sampling_rate/2.5)
    
    # 計算 R 峰之間的間隔（以秒為單位）
    rr_intervals = np.diff(peaks) / sampling_rate
    
    # 計算 BPM
    bpm = 60 / np.mean(rr_intervals)
    
    return bpm

def fit_transform(X, y=None, **fit_params):
        """
        Fit to data, then transform it.

        Fits transformer to `X` and `y` with optional parameters `fit_params`
        and returns a transformed version of `X`.

        Parameters
        ----------
        X : array-like of shape (n_samples, n_features)
            Input samples.

        y :  array-like of shape (n_samples,) or (n_samples, n_outputs), \
                default=None
            Target values (None for unsupervised transformations).

        **fit_params : dict
            Additional fit parameters.

        Returns
        -------
        X_new : ndarray array of shape (n_samples, n_features_new)
            Transformed array.
        """
        # non-optimized default implementation; override when a better
        # method is possible for a given clustering algorithm
        # if y is None:
        #     # fit method of arity 1 (unsupervised transformation)
        #     return fit(X, **fit_params).transform(X)
        # else:
        #     # fit method of arity 2 (supervised transformation)
        #     return fit(X, y, **fit_params).transform(X)

def smooth_signal(listTemp):
        #平滑化
        #用.copy()複製一份listTemp_處理前，以便後續畫圖比較
        listTemp_處理前 = listTemp.copy()
        
        # window_length 控制平滑範圍大小，數值越大，平滑效果越強。
        # polyorder 決定多項式的擬合複雜度，常用低階（如 2 或 3）。
        # mode 用於選擇邊界處理方式，默認 'interp'。
        listTemp[:,0] = scipy.signal.savgol_filter(listTemp[:,0].astype(float), 100, 2)

        #將信號畫成圖型 
        兩圖合併(listTemp_處理前,listTemp,"平滑化")
        return listTemp

#峰度
# 峰度除了正常的常態峰(藍色)，還可以分為高峽峰(紅色)與低闊峰(綠色)。
# 根據Fisher定義，峰度 = 0為常態峰，峰度 > 0為高峽峰，峰度 < 0為低闊峰。
# 若根據Pearson的定義，峰度 = 3為常態峰，峰度 > 3為高峽峰，峰度 < 3為低闊峰。
def Kurtosis(data):
    #     Parameters:
    # a
    # array
    # Data for which the kurtosis is calculated.

    # axis
    # int or None, default: 0
    # If an int, the axis of the input along which to compute the statistic. The statistic of each axis-slice (e.g. row) of the input will appear in a corresponding element of the output. If None, the input will be raveled before computing the statistic.

    # fisher
    # bool, optional
    # If True, Fisher’s definition is used (normal ==> 0.0). If False, Pearson’s definition is used (normal ==> 3.0).

    # bias
    # bool, optional
    # If False, then the calculations are corrected for statistical bias.

    # nan_policy
    # {‘propagate’, ‘omit’, ‘raise’}
    # Defines how to handle input NaNs.

    # propagate: if a NaN is present in the axis slice (e.g. row) along which the statistic is computed, the corresponding entry of the output will be NaN.

    # omit: NaNs will be omitted when performing the calculation. If insufficient data remains in the axis slice along which the statistic is computed, the corresponding entry of the output will be NaN.

    # raise: if a NaN is present, a ValueError will be raised.

    # keepdims
    # bool, default: False
    # If this is set to True, the axes which are reduced are left in the result as dimensions with size one. With this option, the result will broadcast correctly against the input array.

    # Returns
    # :
    # kurtosis
    # array
    # The kurtosis of values along an axis, returning NaN where all values are equal.
    #回傳數值就好
    return scipy.stats.kurtosis(data, axis=0, fisher=True, bias=False)
    # return scipy.stats.kurtosistest(data)

#bias = False意思是消除偏差，同stats.skew()的用法與定義。而fisher = True就是以Fisher的定義，計算並輸出峰度值。
def skewness(data):
    return scipy.stats.skew(data, bias=False)

#TEO（Teager 能量算子）是一种能够有效提取信号能量的非线性算子，对于给定信
#号，TEO 运算能够反映出能量的瞬时变化：
def Teager_power_function(Signal):
    Teager_power = np.zeros(len(Signal))
    # 离散Teager能量算子的公式=S(n)*S(n)-S(n+1)*S(n-1)
    for i in range(1, len(Signal) - 1):
        Teager_power[i] = Signal[i] * Signal[i] - Signal[i + 1] * Signal[i - 1]
    return Teager_power

#計算信號的熵
def entropy(data, base=None):
    """
    計算數據的熵。
    :param data: np.array, 數據序列
    :param base: float, 熵的對數基數（默認為自然對數）
    :return: float, 數據的熵
    """
    _, counts = np.unique(data, return_counts=True)
    return scipy.stats.entropy(counts, base=base)

#計算信號的光譜熵
def spectral_entropy(signal, sampling_rate, base=None):
    """
    計算信號的光譜熵。
    :param signal: np.array, 信號數據
    :param sampling_rate: int, 取樣率
    :param base: float, 熵的對數基數（默認為自然對數）
    :return: float, 信號的光譜熵
    """
    # 計算信號的功率譜密度
    freqs, psd = scipy.signal.welch(signal, fs=sampling_rate)
    
    # 正規化功率譜密度
    psd_norm = psd / np.sum(psd)
    
    # 計算光譜熵
    spectral_entropy_value = scipy.stats.entropy(psd_norm, base=base)
    
    return spectral_entropy_value

def fourier_transform(signal, sampling_rate):
    """
    將信號從時域轉換到頻域。
    :param signal: np.array, 信號數據
    :param sampling_rate: int, 取樣率
    :return: tuple, 包含頻率和對應的傅立葉變換值
    """
    # 計算傅立葉變換
    freqs = np.fft.fftfreq(len(signal), 1/sampling_rate)
    fft_values = np.fft.fft(signal)
    
    return freqs, fft_values

def plot_fourier_transform(signal, sampling_rate):
    """
    繪製信號的傅立葉變換。
    :param signal: np.array, 信號數據
    :param sampling_rate: int, 取樣率
    """
    freqs, fft_values = fourier_transform(signal, sampling_rate)
    
    plt.plot(freqs, np.abs(fft_values))
    plt.title('Fourier Transform')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('Amplitude')
    plt.show()

def writeExecl(list,fileName):
    # import os
    # os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用
    #弄一個新的陣列放置分割逗號後的資料
    tempList = []
    import openpyxl
    wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
    wb.save(fileName)       # 儲存檔案
    wb = openpyxl.load_workbook(fileName, data_only=True)
    # 檢查是否已經存在名為 "Sheet" 的工作表
    sheet_name = 'Sheet'
    if sheet_name in wb.sheetnames:
        # 選擇已存在的工作表
        ws = wb[sheet_name]
    else:
        # 創建一個新的工作表
        ws = wb.create_sheet(sheet_name)

    s1 = wb['Sheet']                        # 開啟工作表 1
    #將list寫入Excel wb,list中用逗號隔開的每個欄位的資料都會寫入到Excel的每個欄位中
    
    

    for i in range(len(list)):
        #先將list[i]的字串裡的[]還有"和'都去掉
        list[i] = str(list[i]).replace("[","").replace("]","").replace("'","").replace('"','')
        
        # 把list[i]用逗號分割出此行的每個欄位
        tempList = str(list[i]).split(",")
        # 將list[i]的每個欄位寫入Excel的每個欄位
        # s1.cell(row=i+1, column=1, value=str(tempList))

        for j in range(len(tempList)-1):
            s1.cell(row=i+1, column=j+1, value=str(tempList[j]))
    
    wb.save(fileName)

def draw_hist(data):
    #bins是指定直方图條形的個數。bins的預設值是10，那麼直方圖會被劃分為10個條形。alpha是透明度。
    plt.hist(data, bins=30, alpha=0.7, color='blue')
    plt.title('Histogram')
    plt.xlabel('Frequency')
    plt.ylabel('Value')
    plt.show()

def drow_time_series(data,times):
    #data,times是兩個list，分別是y軸和x軸的數據
    plt.plot(times,data, alpha=0.7, color='blue')
    plt.title('Time Series')
    plt.xlabel('times')
    plt.ylabel('data')
    plt.show()
import numpy as np
import matplotlib.pyplot as plt

#繪製數據中最穩定的那段區間
def draw_stable_section(data_series,time_series, window_size=1000):
    """
    繪製數據中最穩定的那段區間。
    :param time_series: np.array, 時間序列
    :param data_series: np.array, 數據序列
    :param window_size: int, 用於計算穩定性的窗口大小
    """
    # 計算每個窗口的標準差
    std_devs = [np.std(data_series[i:i+window_size]) for i in range(len(data_series) - window_size + 1)]
    
    # 找到標準差最小的窗口
    min_std_index = np.argmin(std_devs)
    stable_section = data_series[min_std_index:min_std_index + window_size]
    stable_time = time_series[min_std_index:min_std_index + window_size]
    #顯示穩定區間的數據的最大值、最小值

    #圖像盡量大一點
    plt.figure(figsize=(20, 10))

    # 繪製數據和最穩定的區間
    plt.plot(time_series, data_series, label='Original Data')
    plt.plot(stable_time, stable_section, label='Stable Section', color='red')
    plt.legend()
    plt.title('Stable Section of Data;' + "max:" 
              + str(round(np.max(stable_section), 2))
    +";min:" + str(round(np.min(stable_section), 2)))
    plt.xlabel('Time')
    plt.ylabel('Value')
    plt.show()
    #回傳穩定區間的數據的最大值、最小值，取到小數點後兩位
    return round(np.max(stable_section), 2), round(np.min(stable_section), 2),min_std_index,min_std_index + window_size
    #回傳一個新的陣列，內容是穩定區間的數據
    # stablelistTemp = []
    # stablelistTemp.append(stable_section)
    # # stablelistTemp.append(stable_time.timestamp())
    # stablelistTemp.append([dt.timestamp() for dt in stable_time])
    # # return stablelistTemp

    # return stable_section, stable_time
import pywt

def wavelet_baseline_correction(signal_data, wavelet='db4', level=6):
    coeffs = pywt.wavedec(signal_data, wavelet, level=level)
    coeffs[0] = np.zeros_like(coeffs[0])  # 移除最低頻的分量
    corrected_signal = pywt.waverec(coeffs, wavelet)
    return corrected_signal[:len(signal_data)]  # 截斷回原長度

# 其他函數定義...
def draw_scatter(data):
    plt.scatter(range(len(data)), data, alpha=0.7, color='blue')
    plt.title('Scatter Plot')
    plt.xlabel('Index')
    plt.ylabel('Value')
    plt.show()

def get_rppg_wave(peaks, valleys, 前處理結果):
    """
    從峰值和谷值中提取rPPG波形。
    :param peaks: np.array, 峰值
    :param valleys: np.array, 谷值
    :return: np.array, rPPG波形
    """
    # 將峰值和谷值合併，並按時間排序
    rppg_wave = np.sort(np.concatenate([peaks, valleys]))
    
    # 計算rPPG波形
    rppg_wave = np.array([前處理結果[i] for i in rppg_wave])
    #畫出rPPG波形
    plt.plot(rppg_wave)
    plt.title('rPPG Wave')
    plt.xlabel('Time')
    plt.ylabel('Value')
    plt.show()

    return rppg_wave

def 是否符合兔子耳朵的特徵(single_wave):
    """
    判斷單個波形是否符合兔子耳朵的特徵。
    :param single_wave: np.array, 單個波形
    :return: bool, 是否符合兔子耳朵的特徵
    """
    #1.是否有兩個峰值
    #2.第一個峰值是否比第二個峰值高(暫時不看)
    #3.是否只有一個谷值
    #4.谷值是否在中間
    #5.谷值是否比兩個峰值低

    #找到峰值和谷值
    # peaks, _ = find_peaks(single_wave[:,0])
    # valleys, _ = find_peaks(-single_wave[:,0])
    peaks,valleys = find_peaks_and_valleys_不畫圖(single_wave)
    #1.是否有兩個峰值
    if len(peaks) != 2:
        return False
    #2.第一個峰值是否比第二個峰值高
    # if single_wave[peaks[0]] < single_wave[peaks[1]]:
    #     return False
    #3.是否只有一個谷值
    if len(valleys) != 1:
        return False
    #4.先將peaks和valleys整合在同一個陣列
    peaks_valleys = np.concatenate([peaks, valleys])
    #判斷valleys[0]的index是否為1(判斷valleys位置是否在兩個峰值的中間)
    if np.where(peaks_valleys == valleys[0])[0][0] != 2:
        return False
    
    #5.谷值是否比兩個峰值低
    # if single_wave[valleys[0]] > single_wave[peaks[0]] or single_wave[valleys[0]] > single_wave[peaks[1]]:
    #     return False
    
    # 通過檢查，確認為兔子耳朵
    return True

def 單個兔子耳朵基線校準(single_wave):
    """
    對單個兔子耳朵波形進行基線校準。
    :param single_wave: np.array, 單個兔子耳朵波形
    :return: np.array, 基線校準後的波形
    """
    # 找到峰值和谷值
    peaks, _ = find_peaks(single_wave)
    valleys, _ = find_peaks(-single_wave)
    
    # 找到兩個峰值和一個谷值
    peak1, peak2 = peaks
    valley = valleys[0]
    
    # 基線校準
    baseline = np.linspace(single_wave[peak1], single_wave[peak2], len(single_wave))
    baseline[:valley] = single_wave[:valley]
    baseline[valley:] = single_wave[valley]
    
    baseline_corrected_wave = single_wave - baseline
    return baseline_corrected_wave


"""2025.03.12沛蓁 基線校正(每個時窗去做基線校正)"""
def baseline_removal(listTemp,polynomial_degree,window_size):
    # polynomial_degree = 2  # 只適用於 ModPoly 和 IModPoly
    # window_size = 50  # 設定每 50 個點執行一次基線校正

    for i in range(0, len(listTemp), window_size):
        # sub_array = listTemp[i:i+window_size, 0]  # 取出當前的50個點
        sub_array = listTemp[i:i+window_size]  # 取出當前的50個點
        if len(sub_array) < 2:  # 避免長度太短導致算法錯誤
            continue
        baseObj = BaselineRemoval(sub_array)
        # listTemp[i:i+window_size, 0] = baseObj.ModPoly(polynomial_degree)  # 進行基線校正
        listTemp[i:i+window_size] = baseObj.ModPoly(polynomial_degree)  # 進行基線校正
    return listTemp

# def preProcessing(timeWindow:int,第幾個timeWindow:int,file_path_r:str,區間_from:float,區間_to:float):
def preProcessing(timeWindow:int,第幾個timeWindow:int,listTemp:list):
    #listTemp =[]
    listTemp2 =[] 

    #如果不足設定的timeWindow，則跳出訊息提示
    if len(listTemp) < timeWindow:
        # print("用區間篩選後的點數為"+str(len(listTemp))+"個點,不足設定的timeWindow")
        return None, "用區間篩選後的點數為"+str(len(listTemp))+"個點,不足設定的timeWindow"

    #補點
    # even_times = np.linspace(listTemp[0,1], listTemp[len(listTemp)-1,1], 100)
    # listTemp[:,0] = np.interp(even_times, listTemp[:,1], listTemp[:,0])
    # listTemp[:,0]= np.hamming(100) * listTemp[:,0] 
    # listTemp[:,0] = (listTemp[:,0] - np.mean(listTemp[:,0]))/np.std(listTemp[:,0])
                            
    # listTemp[:,0] = listTemp[:,0]/np.linalg.norm(listTemp[:,0]) 

    #將listTemp轉換為float
    # listTemp =listTemp.astype(float)

    #將ROI1的數字畫成圖型
    #plot_data(listTemp)

    #前處理:巴特沃夫處理器(帶通濾波器) ->  平滑化 -> 基線飄移(校準) -> 歸一化 -> 平滑化
    
    #基線飄移(校準)
    try:
        listTemp_處理前 = listTemp.copy()
        
        listTemp[:,0] = baseline_removal(listTemp[:,0],2,200)

        # 小波去除基線飄移
        # original_signal = listTemp[:,0]
        # listTemp[:,0] = wavelet_baseline_correction(original_signal)
        # signal = listTemp[:,0]
        # coeffs = pywt.wavedec(signal, 'db1')  # 小波分解
        # coeffs[0] = np.zeros_like(coeffs[0])  # 将近似系数置零
        # listTemp[:,0] = pywt.waverec(coeffs, 'db1')  # 重构信号
        
        
        # for i in range(int((len(listTemp[:,0])/100))):
        #     listTemp[i*100:i*100+100,0] = signal.detrend(listTemp[i*100:i*100+100,0])
            
        # for i in range(int((len(listTemp[:,0])/100))):
        #     listTemp[i*100:i*100+100,0] = signal.detrend(listTemp[i*100:i*100+100,0])
            
        # # 找到波谷：通过找到信号的局部最小值（即反转信号后找到峰值）
        # signal = listTemp[:,0]
        # # 找到波谷：通过找到信号的局部最小值（即反转信号后找到峰值）
        # inverted_signal = -signal
        # peaks, _ = find_peaks(inverted_signal)

        # # 获取波谷的值
        # valleys = signal[peaks]

        # # 计算波谷的平均值，作为基准线
        # baseline = np.mean(valleys)

        # # 校正信号：将所有波谷调整至基准线，并带动其他点
        # corrected_signal = signal.copy()

        # for i in range(len(signal)):
        #     # 计算当前点到基准线的差值
        #     diff = signal[i] - baseline
        #     # 调整所有点
        #     corrected_signal[i] -= diff  # 向上调整

        # polynomial_degree=2 #only needed for Modpoly and IModPoly algorithm
        # #for test基線校準
        # baseObj=BaselineRemoval(listTemp[:,0])
        # #baseObj=BaselineRemoval(input_array)
        # #listTemp[:,0] = detrend(listTemp[:,0],100,'linear',500)
        # listTemp[:,0]=baseObj.ModPoly(polynomial_degree)
        # #baseObj=BaselineRemoval(listTemp[:,0])
        # # listTemp[:,0]=baseline_correction(listTemp[:,1],listTemp[:,0], 2)
        # #listTemp[:,0]=baseObj.IModPoly(polynomial_degree)
        # # listTemp[:,0]=baseObj.ZhangFit()
        # #將信號畫成圖型 
        # 兩圖合併(listTemp_處理前,listTemp,"基線飄移(校準)")
    except Exception as e:
        print("基線飄移(校準)失敗" + str(e))

    #平滑化
    try:
        #用.copy()複製一份listTemp_處理前，以便後續畫圖比較
        listTemp_處理前 = listTemp.copy()
        #判斷是一維還是二維
        # print("一維還是二維" + str(listTemp[:,0].ndim))
        
        # window_length 控制平滑範圍大小，數值越大，平滑效果越強。
        # polyorder 決定多項式的擬合複雜度，常用低階（如 2 或 3）。
        # mode 用於選擇邊界處理方式，默認 'interp'。
        listTemp[:,0] = scipy.signal.savgol_filter(listTemp[:,0].astype(float), 10, 3)

        #將信號畫成圖型 
        # 兩圖合併(listTemp_處理前,listTemp,"平滑化")
    except Exception as e:
        print("平滑化失敗" + str(e))


    #帶通濾波
    # try:
    #     listTemp_處理前 = listTemp.copy()
        
    #     #原始程式使用的帶通濾波處理程式碼
    #     #帶通濾波器
    #     # processed_samples=listTemp[:,0]
    #     # # processed = butterworth_bandpass_filter(norm,0.1,3.5,self.fps,order=1)
        
    #     # #算norm??????
    #     # # interpolated = np.interp(even_times, self.times, listTemp[:,0])
    #     # interpolated = np.interp(listTemp[:,1], listTemp[:,1], listTemp[:,0])                                 
    #     # # interpolated = np.hamming(L) * interpolated 
    #     # interpolated = np.hamming(listTemp.shape[0]) * interpolated               
    #     # #標準化
    #     # norm = (interpolated - np.mean(interpolated))/np.std(interpolated)
    #     # #變成單位向量
    #     # norm = interpolated/np.linalg.norm(interpolated) 


    #     # processed = butterworth_bandpass_filter(norm,0.1,3.5,30,order=4)
    #     # # processed_samples = butterworth_bandpass_filter(processed_samples,0.8,3,self.fps,order=4)
    #     # processed_samples = butterworth_bandpass_filter(processed_samples,0.8,3,30,order=4)

    #     # processed=signal.detrend(processed)
    #     # processed_samples=signal.detrend(processed_samples)
    #     # listTemp[:,0]=signal.detrend(processed)
    #     # listTemp[:,0]=signal.detrend(processed_samples)

    #     #2024.12.01找到的使用bandpass Butterworth filter对信号数据进行滤波去噪
    #     # time = np.linspace(0, 0.02, 52)
    #     listTemp[:,0]= bandPass_filter(listTemp[:,0])

        
    #     兩圖合併(listTemp_處理前,listTemp,"帶通濾波處理")
    # except Exception as e:
    #             print("帶通濾波處理失敗" + str(e))

    #基線飄移(校準)
    #baseline_als(listTemp2[:,0].astype(float), 1e5, 0.1)
    #for test基線校準
    # input_array=[10,20,1.5,5,2,9,99,25,47]

    # #歸一化
    # try:
    #     listTemp_處理前 = listTemp.copy()
        
    #     # avg = sum(listTemp[:,0])/len(listTemp[:,0])
    #     # listTemp[:,0] = Z_ScoreNormalization(listTemp[:,0],avg,statistics.stdev(listTemp[:,0]))
    #     # 創建MinMaxScaler實例，預設縮放到[0, 1]範圍
    #     scaler = MinMaxScaler()

    #     # 使用MinMaxScaler對特徵進行擬合和轉換
    #     listTemp[:,0] = scaler.fit_transform(listTemp[:,0])
    #     # listTemp[:,0] = listTemp_處理前[:,1]
    #     兩圖合併(listTemp_處理前,listTemp,"歸一化")
    # except Exception as e:
    #             print("歸一化失敗" + str(e))

    #第二次平滑化
    try:
        listTemp_處理前 = listTemp.copy()
        listTemp[:,0] = scipy.signal.savgol_filter(listTemp[:,0].astype(float), 10,2)
        # 兩圖合併(listTemp_處理前,listTemp,"第二次平滑化")
    except Exception as e:
        print("第二次平滑化失敗" + str(e))

    # #做一階差分
    # try:
    #     listTemp[:,0] = np.diff(listTemp[:,0])
    # except Exception as e:
    #     print("一階差分失敗" + str(e))

    #標出波峰和坡谷
    peaks,valleys = find_peaks_and_valleys(listTemp)

    #將上一步的peaks,valleys放進一個陣列裡並紀錄其為peaks or valleys
    _兔子耳朵_count = 0
    特徵值_list = []
    #foreach valleys中的每個valleys
    for i in range(len(valleys)):
        #設定startPointValue為valleys[i]的值
        startPointIndex = valleys[i]

        #假如i+2 > len(valleys) 跳出迴圈
        if i+2 > len(valleys)-1:
            break

        #設定endPoint為valleys[i+2]的值
        #     #假如這個點和前後點的差距有超過0.1，則設定endPointIndex為這個點
        #     if listTemp[valleys[i+2]] - listTemp[valleys[ii]] > 0.1:
        #         endPointIndex = valleys[ii]
        #         break
        endPointIndex = valleys[i+2]

        #將listTemp中從等於startPointIndex為起點，到等於endPointIndex為終點的值，放入 ＿array兔子耳朵
        _兔子耳朵 = listTemp[startPointIndex:endPointIndex]
        
        #基線飄移(校準)
        try:
            _兔子耳朵_處理前 = _兔子耳朵.copy()
            #測試基線飄移(校準)
            # _兔子耳朵[:,0] = baseline_removal(_兔子耳朵[:,0],2,len(_兔子耳朵))
            _兔子耳朵[:,0] = baseline_removal(_兔子耳朵[:,0],2,30)
            #處理前後比較
            # 兩圖合併(_兔子耳朵_處理前,_兔子耳朵,"兔子耳朵基線飄移(校準)")
        except Exception as e:
            print("基線飄移(校準)失敗" + str(e))

        #_兔子耳朵平滑化
        try:
            _兔子耳朵_處理前 = _兔子耳朵.copy()
            _兔子耳朵[:,0] = scipy.signal.savgol_filter(_兔子耳朵[:,0].astype(float), 10,2)
            # 兩圖合併(_兔子耳朵_處理前,_兔子耳朵,"_兔子耳朵平滑化")
        except Exception as e:
            print("_兔子耳朵平滑化失敗" + str(e))
        
        #檢查是否符合兔子耳朵的特徵
        if 是否符合兔子耳朵的特徵(_兔子耳朵) :
            #如果符合，則將兔子耳朵畫成圖型
            # plot_data(_兔子耳朵,"符合兔子耳朵的特徵")  
            
            #如果是第一個符合兔子耳朵的特徵，則將表頭放入特徵值_list的第一欄
            if _兔子耳朵_count == 0:
                # 將表頭放入特徵值_list的第一欄
                特徵值_list.append(特徵值計算(_兔子耳朵,True))
            
            #將符合兔子耳朵的特徵的兔子耳朵放入特徵值_list中的new row
            特徵值_list.append(特徵值計算(_兔子耳朵,False))
            
            _兔子耳朵_count += 1 
            #plot_data(_兔子耳朵,"符合兔子耳朵的特徵")
        else: #若否，則繼續下一個valleys
            continue

        #對_兔子耳朵做基線校準
        # # 小波去除基線飄移
        # original_signal = _兔子耳朵
        # _兔子耳朵 = wavelet_baseline_correction(original_signal)
        # signal = listTemp[:,0]
        # coeffs = pywt.wavedec(signal, 'db1')  # 小波分解
        # coeffs[0] = np.zeros_like(coeffs[0])  # 将近似系数置零
        # listTemp[:,0] = pywt.waverec(coeffs, 'db1')  # 重构信号
        
        # _兔子耳朵2 = 單個兔子耳朵基線校準(_兔子耳朵)

        # 兩圖合併(_兔子耳朵,_兔子耳朵2,"兔子耳朵基線校準")
    
    
    print("共有" + str(_兔子耳朵_count) + "個符合標準的兔子耳朵")
    # 特徵值_list..append(['TKEO_均值', 'TKEO_變異數', 'TKEO_四分位距', 'TKEO_偏度','心率','峰度(Kurtosis)','偏度(skewness)'])  # 第一欄，表頭
    

    return 特徵值_list,""   
    #跳出程式
    # exit()

        
    


    # 進行DFT（傅立葉轉換）

    # # 生成一個示例的時域信號
    # fs = 10  # 采樣頻率
    # # t = np.arange(0, 1, 1/fs)  # 時間序列
    # # frequency1 = 5  # 信號1的頻率（5 Hz）
    # # amplitude1 = 1  # 信號1的振幅
    # # signal1 = amplitude1 * np.sin(2 * np.pi * frequency1 * t)

    # # frequency2 = 20  # 信號2的頻率（20 Hz）
    # # amplitude2 = 0.5  # 信號2的振幅
    # # signal2 = amplitude2 * np.sin(2 * np.pi * frequency2 * t)

    # # # 合併兩個信號
    # # signal = signal1 + signal2

    # listTemp_處理前 = listTemp.copy()
    # fft_result = np.fft.fft(listTemp[:,0])
    # frequencies = np.fft.fftfreq(len(fft_result), 1/fs)

    # # 繪製頻譜圖
    # plt.figure(figsize=(10, 6))
    # plt.subplot(2, 1, 1)
    # plt.plot(t, listTemp[:,0])
    # plt.title("時域信號")

    # plt.subplot(2, 1, 2)
    # plt.plot(frequencies, np.abs(fft_result))
    # plt.title("頻域信號")
    # plt.xlabel("頻率 (Hz)")

    # plt.tight_layout()
    # plt.show()


    #擷取0開始，前面捨棄
    # try:
    #     listTemp_處理前 = listTemp.copy()
        
    #     for i in range(listTemp): 
    #         baseline_values = baseline_als(input_data[i, :]) 

    #     listTemp[:,1] = listTemp_處理前[:,1]
    #     兩圖合併(listTemp_處理前,listTemp,"擷取0開始")
    # except Exception as e:
    #             print("擷取0開始失敗" + str(e))


    # 計算SDNN
    # 需要的input是nni(NN-intervals)，並不是我們原本的signal，所以要先將nni運算出來，這邊就可以先用BiosPPy的ecg()得到
    # R-Peaks indices的值。
    # Load sample ECG signal stored in an OpenSignals file
    #signal = OpenSignalsReader('SampleECG.txt').signal('ECG')

    # Get R-peak locations (and hide the ECG plots)
    # rpeaks = biosppy.signals.ecg.ecg(listTemp[:,0], show=False)[2]

    # # Compute NNI parameters
    # nni = tools.nn_intervals(rpeaks)
    
    # # 但是這邊有一點是必須注意的，就是上面的資料Sample Rate預設是1000，而我的資料集的Sample Rate是2048，所以最後一行要修正成
    # # nni = tools.nn_intervals(rpeaks) / 2.048

    # #parameter nni : array / NN-intervals (ms or s)

    # SDNN = pyhrv.time_domain.sdnn(nni)['sdnn']
    # lnLF = pyhrv.frequency_domain.welch_psd(nni = nni,
    #                             mode='dev')[0]['fft_log'][1]
     #當mode = 'dev'的時候就不會畫圖，共有三種mode，可依需求去選擇
    # return listTemp , ""

def find_peaks_and_valleys(signal_data):
    #畫圖時用時間欄位
    time = signal_data[:,2]
    #畫圖時用index欄位
    # time = signal_data[:,1]
    
    
    signal_data = signal_data[:,0]
    
    time = np.array(time)
    signal_data = np.array(signal_data)

    # 尋找波峰 (peaks)
    peaks, _ = signal.find_peaks(signal_data,distance=5)

    # 尋找波谷 (valleys)（透過反轉訊號後尋找peaks）
    valleys, _ = signal.find_peaks(-signal_data,distance=5)

    # 確保 peaks valleys 是整數陣列
    valleys = valleys.astype(int)
    peaks = peaks.astype(int)
    # valleys = np.array(valleys, dtype=int)
    # peaks = np.array(peaks, dtype=int)
    # print("valleys:", valleys[:5], valleys.dtype)
    # print("time shape:", time.shape, "signal shape:", signal_data.shape)


    # 將波峰與波谷的位置記錄在陣列中
    # peak_points = np.column_stack((time[peaks], signal_data[peaks]))
    # valley_points = np.column_stack((time[valleys], signal_data[valleys]))

    #微軟正黑體
    plt.rc('font', family='Heiti TC')
    # 繪圖標示波峰和波谷
    plt.figure(figsize=(20, 10))
    plt.plot(time, signal_data, label='原始訊號')

    plt.plot(time[valleys], signal_data[valleys], "go", label='波谷')
    plt.plot(time[peaks], signal_data[peaks], "ro", label='波峰')
    
    #x軸的顯示範圍從50開始
    # plt.xlim(0, 500)
    
    plt.xlabel('index')
    plt.ylabel('訊號振幅')
    plt.title('訊號波峰與波谷識別')
    plt.legend()
    plt.grid(True)
    plt.show()
    return peaks,valleys
    # 顯示波峰與波谷點
    # print("波峰位置 (時間, 振幅):")
    # print(peak_points)

    # print("\n波谷位置 (時間, 振幅):")
    # print(valley_points)
def find_peaks_and_valleys_不畫圖(signal_data):
    #畫圖時用時間欄位
    # time = signal_data[:,2]
    #畫圖時用index欄位
    time = signal_data[:,1]

    #放訊號欄位
    signal_data = signal_data[:,0]
    
    time = np.array(time)
    signal_data = np.array(signal_data)

    # 尋找波峰 (peaks)
    peaks, _ = signal.find_peaks(signal_data)

    # 尋找波谷 (valleys)（透過反轉訊號後尋找peaks）
    valleys, _ = signal.find_peaks(-signal_data)

    # 確保 peaks valleys 是整數陣列
    valleys = valleys.astype(int)
    peaks = peaks.astype(int)
    
    return peaks,valleys
 

def 特徵值計算(listTemp,是否為表頭):
    #創建一個list，用來存放特徵值或表頭
    data = []
    
    if 是否為表頭 == True:
        # 將以下要運算特徵值說明作為表頭放入特徵值_list的第一欄
        data.append(['TKEO_均值', 'TKEO_變異數', 'TKEO_四分位距', 'TKEO_偏度','心率'
                     ,'峰度(Kurtosis)','偏度(skewness)','entropy','光譜熵_均值'
                     ,'光譜熵_四分位距','光譜熵_偏度'
                     ,'收縮壓的峰值','舒張壓的峰值','峰值Delt','香農熵'
                     ,'峰值/均方根值','test'])  # 第一欄，表頭
        return data
    
    #1.Teager-Kaiser 能量算子(TKEO)的均值、變異數、四分位距、偏度共4個特徵
    TeagerKaiser = Teager_power_function(listTemp[:,0])

    #Teager-Kaiser均值,四捨五入到小數點後第三位
    TKEO_mean = round(np.mean(TeagerKaiser),3)
    
    #Teager-Kaiser變異數
    TKEO_variance = round(np.var(TeagerKaiser),3)
    #Teager-Kaiser四分位距
    q1 = np.percentile(TeagerKaiser, 25)
    q3 = np.percentile(TeagerKaiser, 75)
    TKEO_iqr = round((q3 - q1),3)
    #Teager-Kaiser偏度
    TKEO_skewness = round(stats.skew(TeagerKaiser),3)

    #2.心率
    #心率
    bpm = round(calculate_bpm(listTemp[:,0],15),3)
    
    #3.峰度(Kurtosis)
    #Kurtosis是一種統計量，用於描述數據分佈的形狀，可以告訴我們數據集中是否有很多極端值（遠離均值的值）
    
    # 確保 listTemp[:, 0] 是數值類型
    numeric_data = listTemp[:, 0].astype(float)
    
    _Kurtosis = round(Kurtosis(numeric_data),3)

    #4.偏度
    #偏度是一種統計量，用於描述數據分佈的對稱性
    # 計算偏度
    _skewness = round(skewness(numeric_data),3)

    #5.entropy
    #熵是一種統計量，用於描述數據的不確定性
    # 計算熵
    _entropy = round(entropy(numeric_data),3)

    #6. 
    #光譜熵(=頻譜熵)
    # 計算光證熵
    # _spectral_entropy = round(spectral_entropy(numeric_data, 15),3)
    # FFT 設定
    L_FFT = 512
    Xn = np.fft.fft(numeric_data, n=L_FFT)
    # freqs = np.fft.fftfreq(L_FFT, 1/Fs)
    #對Xn作正規化
    Xn = Xn / np.linalg.norm(Xn)
    # 對Xn求熵
    #對Xn的每一個值都求熵
    for i in range(len(Xn)):
        #對Xn的每一個值都求熵
        Xn[i] = Xn[i] * np.log(Xn[i])
    #計算Xn的均值、四分位距、偏度
    Xn_mean = round(np.mean(Xn),3)
    Xn_iqr = round(np.percentile(Xn, 75) - np.percentile(Xn, 25),3)
    Xn_skewness = stats.skew(Xn)
    # Xn_kurtosis = stats.kurtosis(Xn)

    

    # 頻(光)譜熵（Spectral Entropy）是一種衡量訊號在頻率域中能量分布的複雜性與隨機性的指標，常用在 RPPG、EEG、ECG 等生理訊號中。
    # 🔹 概念簡介
    # 頻譜熵會將訊號進行 傅立葉轉換（FFT），得到不同頻率上的能量分佈（功率譜）。
    # 然後將這些能量正規化為機率分布，再使用Shannon entropy公式計算訊號在頻域的「混亂程度」。
    # 🔸 數值意義（越高越雜亂）：
    # 頻譜熵數值	意義
    # 接近 0	頻譜能量集中（單一或穩定週期）
    # 趨近最大值	能量分布廣泛（隨機、無週期性）
    
    # 7.收縮壓和舒張的峰值及Delt(收縮壓峰值時間和舒張壓的峰值時間的差)
    
    #標出波峰和坡谷
    # peaks,valleys = find_peaks_and_valleys(listTemp)
    peaks,valleys = find_peaks_and_valleys_不畫圖(listTemp)
    
    #收縮壓峰值 #預設為第一個波峰
    # systolic_peak = np.max(listTemp)
    systolic_peak = round(listTemp[:,0][peaks[0]],3)
    #舒張壓峰值 #預設為第二個波峰
    # diastolic_peak = np.min(numeric_data)
    diastolic_peak = round(listTemp[:,0][peaks[1]],3)

    #產生一個空的timestamp
    # 初始化一個最小的時間戳
    timestamp = datetime.min
    
    #將listTemp[:,2][peaks[0]]轉為timestamp
    systolic_peak_time_1 = datetime.fromtimestamp(listTemp[:,2][peaks[0]])
    systolic_peak_time_2 = datetime.fromtimestamp(listTemp[:,2][peaks[1]])

    #Delt(收縮壓峰值時間和舒張壓的峰值時間的差)
    Delt = systolic_peak_time_1 - systolic_peak_time_2
    Delt_show = str(Delt.microseconds)

    #8.香農熵 (Entropy)Entropy=−t∑t=1p(S(t))log2 p(S(t)
    # 熵是一種統計量，用於描述數據的不確定性
    
    # 統計每個值的出現次數並轉為機率
    values, counts = np.unique(listTemp[:,0], return_counts=True)
    probabilities = counts / counts.sum()

    # 計算香農熵
    Shannon_Entropy = round(-np.sum(probabilities * np.log2(probabilities)),3)

    #9.Peak to RMS（峰值/均方根值）•Peak to RMS比值越大：表示信號的瞬時最大值遠高於其平
    # 均能量水平，具有較大的動態範圍。
    # •Peak to RMS比值越小：表示信號的瞬時最大值接近其平均
    # 能量水平，具有較小的動態範圍。
    #V_peak：信號的峰值。這是信號在一段時間達到的最大值。
    V_peak = np.max(listTemp[:,0])
    #訊號的 RMS（Root Mean Square）均方根值 常見於訊號能量、電壓、聲音強度等分析中。
    V_RMS = np.sqrt(np.sum(listTemp[:,0]**2)/len(listTemp))
    Peak_to_RMS = round((V_peak / V_RMS),3)

    #?.
    #傅立葉變換
    # 繪製傅立葉變換
    #plot_fourier_transform(numeric_data, 15)

    
    # data.append(['TKEO_均值', 'TKEO_變異數', 'TKEO_四分位距', 'TKEO_偏度','心率','峰度(Kurtosis)','偏度(skewness)'])  # 第一欄，表頭
    #創建表頭，包含上面的所有特徵

    data.append([ TKEO_mean,TKEO_variance, TKEO_iqr, TKEO_skewness,bpm,_Kurtosis
                 ,_skewness,_entropy,Xn_mean,Xn_iqr,Xn_skewness,systolic_peak,diastolic_peak,
                 Delt_show,Shannon_Entropy,Peak_to_RMS,0])  # 第二欄以後，值
    # data.append([ 111, 113, 114])  # 第二欄以後，值
    # data.append([ 88, 99, 77])  # 第二欄以後，值

    return data

# if __name__ == '__main__':
    #start!
    #宣告變數
    # face_frame = np.zeros((10, 10, 3), np.uint8)
    # mask = np.zeros((10, 10, 3), np.uint8)
    
    # #臉頰
    # ROI1 = np.zeros((10, 10, 3), np.uint8)
    # ROI2 = np.zeros((10, 10, 3), np.uint8)

    
    # #下巴
    # ROI3 = np.zeros((10, 10, 3), np.uint8)
    # ROI4 = np.zeros((10, 10, 3), np.uint8)
    
    # #全臉
    # ROI5 = np.zeros((10, 10, 3), np.uint8)
    # ROI6 = np.zeros((10, 10, 3), np.uint8)
    
    # #額頭
    # ROI7 = np.zeros((10, 10, 3), np.uint8)
    # ROI8 = np.zeros((10, 10, 3), np.uint8)
    
    # leftEye = np.zeros((6, 2), np.uint8)
    # rightEye = np.zeros((6, 2), np.uint8)
    # status = False
    # roi = _idxROI  #選擇ROI 0:全臉 1:額頭 2:下巴 3:臉頰

    #plot_data(numbers_list)
    
    
